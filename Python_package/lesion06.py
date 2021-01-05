import openpyxl
# wb =openpyxl.load_workbook('test_case_api.xlsx')#这里openpyxl是一个函数，前面加一个变量
# sheet = wb['register']
# row_max = sheet.max_row
# # print(row_max)
# case_list=[0]#新建空列表，存放for循环依次读取到的测试用例数据
# for j in range(2,row_max+1):
#     data_dict =dict(
#     cell_id=sheet.cell(row=j, column=1).value,
#     url=sheet.cell(row=j, column=5).value,
#     data=sheet.cell(row=j, column=6).value,
#     expect=sheet.cell(row=j, column=7).value
#     )
#     case_list.append(data_dict)#把每一行读取到的测试用例数据生成的字典，追加到list中
# print(case_list)
#可能不止读取一个sheet表或者文件名，读取Excel用例，需要封装成一个函数
def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename)#这里openpyxl是一个函数，前面加一个变量
    sheet = wb[sheetname]
    row_max = sheet.max_row
    case_list=[]#新建空列表，存放for循环依次读取到的测试用例数据
    for j in range(2,row_max+1):
        data_dict =dict(
        cell_id=sheet.cell(row=j, column=1).value,
        url=sheet.cell(row=j, column=5).value,
        data=sheet.cell(row=j, column=6).value,
        expect=sheet.cell(row=j, column=7).value
        )
        case_list.append(data_dict) #把每一行读取到的测试用例数据生成的字典，追加到list中
    return case_list
#调用函数
result=read_case('../test_data/test_case_api.xlsx', 'login')# 函数调用的返回值，赋给变量
print(result)
# #读写结果数据到Excel表中
# wb = openpyxl.load_workbook('test_case_api.xlsx')#这里openpyxl是一个函数，前面加一个变量
# sheet = wb['login']
# sheet.cell(row=2,column=8).value=('passed')
# wb.save('test_case_api.xlsx')
# def write_result(filename,sheetname,row,column,finnalresult):
#     wb = openpyxl.load_workbook(filename)  # 这里openpyxl是一个函数，前面加一个变量
#     sheet = wb[sheetname]
#     sheet.cell(row='row',column='column').value='finnalresult'
#     wb.save(filename)
#     return filename
# result_write=write_result('test_case_api.xlsx','login')
# print(result_write)
# tuple=(1,)
# print(tuple)
# print(type(tuple))
# ss=(1)
# print(ss)
# print(type(ss))