import openpyxl
import requests
#读取Excel中测试用例数据，封装成一个函数
# def login(url,data):
#     headers_log ={'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}
#     response_log =requests.post(url=url,headers=headers_log,json=data).json()
#     #返回值
#     return response_log
#     #调用login（）
# url_log = 'http://120.78.128.25:8766/futureloan/member/login'
# data_log = {"mobile_phone": "15678328798", "pwd": "lemon666"}
# result=login(url=url_log,data=data_log)
# print(result)
'''
接口自动化测试
1)excel测试用例准备OK，代码可以自动读取用例数据-----read_case(filename,sheetname)
2）执行接口测试，得到响应结果---- login(url,data)
3）断言：响应结果=====预期结果-----通过/不通过
4）写到最终执行通过与否的结果---Excel表格---------write_result(filename,sheetname,row,column,finnalresult)
'''

# 1)excel测试用例准备OK，代码可以自动读取用例数据-
def read_case(filename,sheetname):
    wb = openpyxl.load_workbook(filename)#这里openpyxl是一个函数，前面加一个变量
    sheet = wb[sheetname]
    row_max = sheet.max_row
    case_list=[]#新建空列表，存放for循环依次读取到的测试用例数据
    for j in range(2,row_max+1):
        data_dict =dict(
        cell_id=sheet.cell(row=j, column=1).value,
        url=sheet.cell(row=j, column=5).value,
        data=sheet.cell(row=j, column=6).value,
        expect=sheet.cell(row=j, column=7).value
        )
        case_list.append(data_dict) #把每一行读取到的测试用例数据生成的字典，追加到list中
    return case_list
# result=read_case('test_case_api.xlsx','login')# 函数调用的返回值，赋给变量
# print(result)
# 2）执行接口测试，得到响应结果
def login(url,data):
    headers ={'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}
    response=requests.post(url=url,headers=headers,json=data).json()
    return response
# url_log = 'http://120.78.128.25:8766/futureloan/member/login'
# data_log = {"mobile_phone": "15678328798", "pwd": "lemon666"}
# result=login(url=url_log,data=data_log)
# print(result)
# #4)写到最终执行通过与否的结果---Excel表格
wb = openpyxl.load_workbook('../test_data/test_case_api.xlsx')#这里openpyxl是一个函数，前面加一个变量
sheet = wb['login']
sheet.cell(row=2,column=8).value=('passed')
wb.save('test_case_api.xlsx')
def write_result(filename,sheetname,row,column,finnalresult):
    wb = openpyxl.load_workbook(filename)  # 这里openpyxl是一个函数，前面加一个变量
    sheet = wb[sheetname]
    sheet.cell(row=row,column=column).value =finnalresult
    wb.save(filename)
##3)断言 实际结果=预期结果

# eval()---运行字符串包裹着的表达式，包括元祖、列表、字典
# data = eval('{"mobile_phone":"15512345678","pwd":"12345678"}')
# print(data)
# print(type(data))
# str1 = '2+6'
# print(eval(str1))
